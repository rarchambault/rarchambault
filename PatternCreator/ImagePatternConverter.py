from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from PIL import Image

def image_to_excel(image_path, excel_path, n_rows, n_cols, n_colors, ignore_color=None):
    # Load the image
    img = Image.open(image_path)
    img = img.convert('RGB')  # Convert to RGB if not already
    
    # Resize the image to the desired grid size
    img_resized = img.resize((n_cols, n_rows), Image.Resampling.LANCZOS)
    
    # Quantize the image to reduce the number of colors
    img_quantized = img_resized.quantize(colors=n_colors, method=1)
    
    # Convert the image back to RGB (since quantize returns a palette-based image)
    img_quantized = img_quantized.convert('RGB')
    
    # Get pixel data
    pixels = img_quantized.load()

    # Convert ignore_color to RGB if provided (accepts hex or tuple)
    if ignore_color:
        if isinstance(ignore_color, str):
            ignore_color = tuple(int(ignore_color[i:i+2], 16) for i in (0, 2, 4))  # Convert hex to RGB
    
    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Function to check if the color matches the ignore_color
    def is_ignored_color(r, g, b):
        if ignore_color:
            return (r, g, b) == ignore_color
        return False
    
    # Loop through the quantized pixels and set Excel cell background colors
    for row in range(1, n_rows + 1):
        for col in range(1, n_cols + 1):
            # Get the RGB values of the quantized pixel
            r, g, b = pixels[col - 1, row - 1]
            
            # Skip if this pixel matches the ignored color
            if is_ignored_color(r, g, b):
                continue  # Skip the current cell
            
            # Convert RGB to hex
            hex_color = f'{r:02X}{g:02X}{b:02X}'
            
            # Set the cell background color
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.value = ""  # Set cell value as empty
            
            # Adjust column width and row height
            ws.column_dimensions[get_column_letter(col)].width = 2.5
        ws.row_dimensions[row].height = 15
    
    # Save the Excel file
    wb.save(excel_path)

# Example usage
image_path = "IMG_9655.PNG"  # Path to the input image
excel_path = "output_pattern.xlsx"  # Path where the Excel file will be saved
n_rows, n_cols = 110, 60  # Grid dimensions
n_colors = 20  # Number of colors to reduce to (you can change this)
ignore_color = None  # Optional color to ignore (white in this case)

image_to_excel(image_path, excel_path, n_rows, n_cols, n_colors, ignore_color)
